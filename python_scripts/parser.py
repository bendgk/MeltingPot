import re
import pprint
from openpyxl import load_workbook
import json
from PIL import Image, ImageDraw

pp = pprint.PrettyPrinter(indent=4)

wb = load_workbook('data.xlsx')
ws = wb.active

ingredients = {}
recipes = {}

for col in [1, 2, 3, 4, 6, 7, 9]:
    category = ""
    
    for row in range(2, 15):
        cell_data = str(ws.cell(column=col, row=row).value).strip().lower()
        if row == 2: category = cell_data
        
        elif cell_data != "none":
            match = re.match(r"[a-zA-Z\/]\D*", cell_data)
            
            if match:
                if category in ingredients:
                    ingredients[category].append([match.group(0)[0:-1], 0])
                else:
                    ingredients[category] = [[match.group(0)[0:-1], 0]]
                    
for col in range(2, 27):
    culture = ""
    recipe = ""
    
    for row in range(17, 40):
        cell_data = str(ws.cell(column=col, row=row).value).strip().lower()
        
        if row == 17:
            culture = cell_data.strip()
        elif row == 18:
            recipe = cell_data.strip()
            recipes[recipe] = {'culture': culture, 'ingredients': []}
            
        elif cell_data != "none":
            if "(o)" in cell_data:
                #order is recipes[rercipe] = [[Ingredient, Required], [Ingredient, Required]]
                #where Ingredient is an ingredient and Required is a boolean
                recipes[recipe]['ingredients'].append([cell_data.replace("(o)", "").strip(), False])
            else:
                recipes[recipe]['ingredients'].append([cell_data.strip(), True])
                
            for key, value in ingredients.items():
                for food in value:
                    if food[0] in cell_data:
                        food[1] += 1

with open('ingredients.json', 'w+') as f:
    json.dump(ingredients, f, indent=4)


with open('recipes.json', 'w+') as f:
    json.dump(recipes, f, indent=4)
    
for value in ingredients.values():
    for ingredient in value:
        ingredient = ingredient[0].replace(" ", "_")
        ingredient = ingredient.replace("/", "-")
        
        print(ingredient)
        
        img = Image.new('RGB', (200, 200), color = (0, 0, 0))
        d = ImageDraw.Draw(img)
        d.text((100,100), ingredient, fill=(255,255,255))
        img.save('images/' + ingredient + ".jpg")
